# app/utils/excel_parser.py
from io import BytesIO

from openpyxl import load_workbook

from app.schemas.quiz_import import (
    ParsedAnswerData,
    ParsedQuestionData,
    ParsedQuizData,
    QuizValidationError,
)

EXPECTED_HEADERS = [
    "quiz_title",
    "quiz_description",
    "question_title",
    "answer_text",
    "is_correct",
]

TRUE_VALUES = {"true", "1", "yes", "так"}


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in TRUE_VALUES
    return False


def parse_excel_to_quizzes(
    file_content: bytes,
) -> tuple[list[ParsedQuizData], list[QuizValidationError]]:
    """
    Читає Excel-файл і групує плоскі рядки у вкладену структуру
    quiz -> questions -> answers.
    Повертає (список квізів, список помилок парсингу по рядках).
    Не кидає виняток через одну погану клітинку — пропускає рядок.
    """
    workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    headers = next(rows, None)
    if headers is None or list(headers) != EXPECTED_HEADERS:
        workbook.close()
        raise ValueError(f"Invalid Excel headers. Expected: {EXPECTED_HEADERS}")

    quizzes: dict[str, ParsedQuizData] = {}
    errors: list[QuizValidationError] = []

    for row_number, row in enumerate(rows, start=2):
        if all(value is None for value in row):
            continue

        try:
            (
                quiz_title_raw,
                quiz_description_raw,
                question_title_raw,
                answer_text_raw,
                is_correct_raw,
            ) = row

            quiz_title = str(quiz_title_raw).strip() if quiz_title_raw is not None else ""
            question_title = str(question_title_raw).strip() if question_title_raw is not None else ""
            answer_text = str(answer_text_raw).strip() if answer_text_raw is not None else ""

            if not quiz_title or not question_title or not answer_text:
                errors.append(
                    QuizValidationError(
                        quiz_title=quiz_title or "unknown",
                        row_number=row_number,
                        message="Quiz title, question title, and answer text must not be empty",
                    )
                )
                continue

            if quiz_title not in quizzes:
                quizzes[quiz_title] = ParsedQuizData(
                    title=quiz_title,
                    description=str(quiz_description_raw or "").strip(),
                    questions=[],
                    source_row_start=row_number,
                )

            quiz = quizzes[quiz_title]
            question = next((q for q in quiz.questions if q.title == question_title), None)
            if question is None:
                question = ParsedQuestionData(title=question_title, answers=[])
                quiz.questions.append(question)

            question.answers.append(
                ParsedAnswerData(text=answer_text, is_correct=_parse_bool(is_correct_raw))
            )

        except Exception as exc:  # noqa: BLE001 — навмисно широкий catch на рівні рядка
            errors.append(QuizValidationError(
                quiz_title="unknown", row_number=row_number, message=f"Failed to parse row: {exc}",
            ))
            continue

    workbook.close()
    return list(quizzes.values()), errors